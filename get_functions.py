import csv 
from openpyxl import Workbook
from openpyxl import load_workbook
from sr_log import sr_log_messages, sr_debug

def create_workbook(wbook):
   
    try:
        wb = load_workbook(wbook)
        print("Workbook '%s'exists" %(wb))
    except:
        print("Creating worksheet: '%s'" %(wbook))
        wb = Workbook()

    wb.save(wbook)

def get_read_csv_files(file_name):
    datalist = []
    fd = open(file_name)

    fname = csv.reader(fd)
    for row in fname:
        datalist.append(row)
    return datalist

def create_worksheet(wbname, dst_wname):

    wb = load_workbook(wbname)


    try:
        dwsheet = wb.get_sheet_by_name(dst_wname)
        sr_debug ("worksheet '%s' found"%(dst_wname))
        sr_debug ("removing worksheet:'%s'"%(dst_wname))
        wb.remove_sheet(dwsheet)
    except:
        sr_debug ("worksheet '%s' not found"%(dst_wname))
    finally:
        sr_debug ("creating new worksheet:'%s'"%(dst_wname))
        dwsheet = wb.create_sheet(dst_wname,0)

    wb.save(wbname)
