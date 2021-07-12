"""Import necessary modules"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.formatting.rule import FormatObject, DataBar, Rule
from get_functions import create_workbook, get_read_csv_files, create_worksheet
from columns import get_column_names_and_positions
from sr_log import sr_debug

(SNO_POS, SNO_COL, ISSUE_KEY_POS, ISSUE_KEY_COL_NAME,
 ISSUE_ID_POS, ISSUE_ID_COL_NAME, CUSTOM_FIELD_EPICLINK_POS,
 CUSTOM_FIELD_EPICLINK_COL_NAME, ENAME_POS, ENAME_COL_NAME,
 ASSIGNEE_POS, ASSIGNEE_COL_NAME, CUSTOM_FIELD_STORYPOINTS_POS,
 CUSTOM_FIELD_STORYPOINTS_COL_NAME, TESTE_POS, TESTE_COL_NAME,
 ORIGINAL_ESTIMATE_POS, ORIGINAL_ESTIMATE_COL_NAME, TIME_SPENT_POS,
 TIME_SPENT_COL_NAME, REMAINING_ESTIMATE_POS, REMAINING_ESTIMATE_COL_NAME,
 SPRINT_POS, SPRINT_COL_NAME, SPRINT2_POS, SPRINT2_COL_NAME, SPRINT3_POS,
 SPRINT3_COL_NAME, SUMMARY_POS, SUMMARY_COL_NAME, EPICS_CUSTOMFIELD_POS,
 EPICS_ESDATE_POS, EPICS_ETDATE_POS, EPICS_EASDATE_POS, EPICS_EAEDATE_POS,
 PROGRESS_POS, PROGRESS_GEN_COL_NAME, SCHEDULED_PROGRESS_POS,
 SCHEDULED_PROGRESS_GEN_COL_NAME, SCHEDULED_OVERRUN_POS,
 SCHEDULED_OVERRUN_GEN_COL_NAME, REMARKS_POS,
 REMARKS_GEN_COL_NAME) = get_column_names_and_positions()


def get_col_names(wbname, dst_wname):
    """This function defined to give names for columns in excel sheet"""
    wb_name = load_workbook(wbname)
    dwsheet = wb_name[dst_wname]
    dwsheet.insert_cols(1, 18)
    dwsheet[1][SNO_POS].value = SNO_COL
    dwsheet[1][ISSUE_KEY_POS].value = ISSUE_KEY_COL_NAME
    dwsheet[1][ISSUE_ID_POS].value = ISSUE_ID_COL_NAME
    dwsheet[1][CUSTOM_FIELD_EPICLINK_POS].value = CUSTOM_FIELD_EPICLINK_COL_NAME
    dwsheet[1][ENAME_POS].value = ENAME_COL_NAME
    dwsheet[1][ASSIGNEE_POS].value = ASSIGNEE_COL_NAME
    dwsheet[1][CUSTOM_FIELD_STORYPOINTS_POS].value = CUSTOM_FIELD_STORYPOINTS_COL_NAME
    dwsheet[1][TESTE_POS].value = TESTE_COL_NAME
    dwsheet[1][ORIGINAL_ESTIMATE_POS].value = ORIGINAL_ESTIMATE_COL_NAME
    dwsheet[1][TIME_SPENT_POS].value = TIME_SPENT_COL_NAME
    dwsheet[1][REMAINING_ESTIMATE_POS].value = REMAINING_ESTIMATE_COL_NAME
    dwsheet[1][SPRINT_POS].value = SPRINT_COL_NAME
    dwsheet[1][SPRINT2_POS].value = SPRINT2_COL_NAME
    dwsheet[1][SPRINT3_POS].value = SPRINT3_COL_NAME
    dwsheet[1][SUMMARY_POS].value = SUMMARY_COL_NAME
    dwsheet['p1'] = PROGRESS_GEN_COL_NAME
    dwsheet['q1'] = SCHEDULED_PROGRESS_GEN_COL_NAME
    dwsheet['r1'] = SCHEDULED_OVERRUN_GEN_COL_NAME
    dwsheet['s1'] = REMARKS_GEN_COL_NAME

    wb_name.save(wbname)

def get_sprint_value(wbname, stories, dst_wname):
    """Function to compare sprint data, sort them and append it to list"""
    sprint_list = []
    wb_name = load_workbook(wbname)
    dwsheet = wb_name[dst_wname]
    srcount1 = len(stories)
    for j in range(1, srcount1):
        sprint_list.append(stories[j][SPRINT_POS])
        sprint_list.append(stories[j][SPRINT2_POS])
        sprint_list.append(stories[j][SPRINT3_POS])
        sprint_list = list(filter(None, sprint_list))
        sprint_list.sort(reverse=True)
        dwsheet[j + 1][ISSUE_ID_POS].value = sprint_list[0]
        sprint_list = []
    wb_name.save(wbname)


def get_values_for_columns(wbname, stories, dst_wname):
    """This function defined to give values for columns in excel sheet"""
    wb_name = load_workbook(wbname)
    dwsheet = wb_name[dst_wname]
    srcount1 = len(stories)
    sr_debug("stories length: %d" % (srcount1))
    for i in range(1, srcount1):
        dwsheet[i + 1][SNO_POS].value = int(stories[i][SNO_POS])
        dwsheet[i + 1][ISSUE_KEY_POS].value = stories[i][
            CUSTOM_FIELD_EPICLINK_POS]
        dwsheet[i + 1][CUSTOM_FIELD_EPICLINK_POS].value = stories[i][
            ISSUE_KEY_POS]
        dwsheet[i + 1][ENAME_POS].value = stories[i][SUMMARY_POS]
        dwsheet[i + 1][ASSIGNEE_POS].value = stories[i][ASSIGNEE_POS]
        dwsheet[i+1][CUSTOM_FIELD_STORYPOINTS_POS].value = int(stories[i][
            CUSTOM_FIELD_STORYPOINTS_POS])
        estimate_effort_hours = int(stories[i][ORIGINAL_ESTIMATE_POS]) / 3600
        dwsheet[i+1][TIME_SPENT_POS].value = (estimate_effort_hours)
        dwsheet[i+1][REMAINING_ESTIMATE_POS].value = int(stories[i]
                                                         [TIME_SPENT_POS])

        scheduled_progress = 100
        dwsheet[i+1][SCHEDULED_PROGRESS_POS].value = ('{}{}'.format
                                                      (scheduled_progress,
                                                       "%"))

        scheduled_overrun = 0
        dwsheet[i+1][SCHEDULED_OVERRUN_POS].value = ('{}{}'.format
                                                     (scheduled_overrun, "%"))

    wb_name.save(wbname)


def get_dates_append(wbname, stories, epics, dst_wname):
    """Function to compare custom epic link, give particular dates"""
    dates_list = []

    wb_name = load_workbook(wbname)
    dwsheet = wb_name[dst_wname]

    srcount1 = len(stories)
    sr_debug("stories length: %d" % (srcount1))

    srcount2 = len(epics)
    sr_debug("epics length: %d" % (srcount2))
    for k in range(1, srcount2):
        for i in range(1, srcount1):
            dates_list.append(epics[k][EPICS_ESDATE_POS])
            dates_list.append(epics[k][EPICS_ETDATE_POS])
            dates_list.append(epics[k][EPICS_EASDATE_POS])
            dates_list.append(epics[k][EPICS_EAEDATE_POS])
            if epics[k][EPICS_CUSTOMFIELD_POS] == stories[i][
                    CUSTOM_FIELD_EPICLINK_POS]:
                dwsheet[i + 1][TESTE_POS].value = (dates_list[0])
                dwsheet[i + 1][ORIGINAL_ESTIMATE_POS].value = (dates_list[1])
                dwsheet[i + 1][SPRINT3_POS].value = (dates_list[2])
                dwsheet[i + 1][SUMMARY_POS].value = (dates_list[3])
            dates_list = []

    sr_debug("Saving :%s" % (wbname))
    wb_name.save("employee-details.xlsx")
    wb_name.save(wbname)


def get_cell_alignment(wbname, dst_wname):
    """Function to align the cell values in a particular position"""
    wb_name = load_workbook(wbname)
    dwsheet = wb_name[dst_wname]

    text_alignment_for_column = Alignment(horizontal="center",
                                          vertical="center", wrapText=True)
    text_alignment_for_row = Alignment(horizontal="center", vertical="center")

    row_count = dwsheet.max_row
    column_count = dwsheet.max_column

    for i in range(2, row_count+1):
        for j in range(0, column_count):
            dwsheet[1][j].alignment = text_alignment_for_column
            dwsheet[i][j].alignment = text_alignment_for_row
    wb_name.save(wbname)


def get_cell_colors_using_patternfill(wbname, dst_wname):
    """This function is to get colors for column names"""
    wb_name = load_workbook(wbname)
    dwsheet = wb_name[dst_wname]

    column_count = dwsheet.max_column

    fill_pattern = PatternFill(patternType='solid', fgColor='CCCCFF')
    for i in range(0, column_count):
        dwsheet[1][i].fill = fill_pattern

    wb_name.save(wbname)


def get_heading(wbname, dst_wname):
    """This function is used to insert heading and give color to the heading"""
    wb_name = load_workbook(wbname)
    dwsheet = wb_name[dst_wname]
    dwsheet.insert_rows(1)
    dwsheet['A1'].value = "SPRINT/STORY BOARD"
    dwsheet.merge_cells('A1:S1')
    dwsheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

    fill_pattern = PatternFill(patternType='solid', fgColor='FFFF00')
    dwsheet['A1'].fill = fill_pattern
    wb_name.save(wbname)


def get_cell_formulae(wbname, dst_wname):
    """This function defined to get formulae into the cell"""
    wb_name = load_workbook(wbname)
    dwsheet = wb_name[dst_wname]
    dwsheet['M3'].value = "=J3-L3"
    dwsheet['P3'].value = "=J3/L3*100"
    drcount = dwsheet.max_row
    for i in range(3, drcount + 1):
        dwsheet['L' + str(i)].value = "="+"K"+str(i)+"/"+"3600"
        dwsheet['M' + str(i)].value = "=" + "J" + str(i) + "-" + "L" + str(i)
        dwsheet['P' + str(i)].value = "="+"L"+str(i)+"/"+"J"+str(i)+"%"

    wb_name.save(wbname)

def inserting_data_bar(wbname, dst_wname):
    """This function defined for adding data-bar to the cells"""
    wb_name = load_workbook(wbname)
    dwsheet = wb_name[dst_wname]
    row_count = dwsheet.max_row
    first = FormatObject(type='percent')
    second = FormatObject(type='percent')
    data_bar = DataBar(cfvo=[first, second], color="FF0000", showValue=None,
                         minLength=None, maxLength=None)
    # assign the data bar to a rule
    rule = Rule(type='dataBar', dataBar=data_bar)
    dwsheet.conditional_formatting.add("P3:P" + str(row_count), rule)

    wb_name.save(wbname)

def main():
    """Main function is defined to call all the functions"""
    filename = "employee-details.xlsx"
    stories = "02-stories.csv"
    epics = "01-epics.csv"
    dst_wname = "generated"
    create_workbook(filename)
    epics_data = get_read_csv_files(epics)
    stories_data = get_read_csv_files(stories)
    create_worksheet(filename, dst_wname)
    get_col_names(filename, dst_wname)
    get_sprint_value(filename, stories_data, dst_wname)
    get_values_for_columns(filename, stories_data, dst_wname)
    get_dates_append(filename, stories_data, epics_data, dst_wname)
    get_cell_alignment(filename, dst_wname)
    get_cell_colors_using_patternfill(filename, dst_wname)
    get_heading(filename, dst_wname)
    get_cell_formulae(filename, dst_wname)
    inserting_data_bar(filename, dst_wname)


if __name__ == '__main__':
    main()
